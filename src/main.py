# 실제 구현 환경

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# GET 요청 (데이터 조회)
resp = requests.get('https://news.naver.com/section/104')

# 상태 코드 확인
print(resp.status_code)  # 200: 성공, 404: 찾을 수 없음, 500: 서버 에러

# 응답 내용 확인
if(resp.status_code == 200):
  # print(resp.url) # url 정보
  # print(resp.json) # json 데이터
  print(resp.text)  # HTML 문서를 가져옴
else:
  print("응답에 실패했습니다.")
  

soup = BeautifulSoup(resp.text, 'html.parser')

# 뉴스기사 제목 검색
news_title = soup.select('.section_latest_article .sa_text_strong')
print(news_title)

news_title_text = []

# 1차 가공 : 뉴스 기사에 제목 텍스트만 추출
for idx, title in enumerate(news_title):  
  title_text = title.get_text(strip=True)
  news_title_text.append(title_text)  
  
# 뉴스 기사 제목 출력
print("== 뉴스 기사 제목 출력 ==")
for i, title in enumerate(news_title_text):
  no = i + 1
  print(f"{no} : {title}")
  
# 특정 키워드로 제목 검색  
# title.find(keyword) -> 결과값이 0아니면 -1
# 0인 경우에 내가 원하는 키워드가 존재한다.
# -1인 경우에는 내가 원하는 키워드가 존재하지 않는다.

print("== 키워드로 뉴스 기사 검색 ==")
find_keyword_list = []
keyword = '트럼프'
for i, title in enumerate(news_title_text):
  no = i + 1    
  
  if title.find(keyword) != -1:
    # print(f"{no} : {title}")
    find_keyword_list.append(title)
    
print(find_keyword_list) # 특정 키워드로 걸러진 데이터가 리스트에 저장

# 특정 키워드를 가진 뉴스 기사 제목을 반복문을 이용한 출력
print("== 키워드 뉴스 제목 리스트 순회 ==")
for i, title in enumerate(find_keyword_list):
  print(f"{i + 1} : {title}")

def adjust_excel_format(filename):
    """엑셀 파일의 행과 열 크기를 조절하는 함수"""
    try:
        # 엑셀 파일 로드
        wb = load_workbook(filename)
        ws = wb.active # 현재 활성화된 시트 선택
        
         # 스타일 정의
        header_font = Font(
          name='맑은 고딕',  # 폰트
          size=16,          # 크기
          bold=True,        # 굵게
          color='FFFFFF'    # 흰색
        )
        
        header_fill = PatternFill(
            start_color='366092', # 파란색 계열
            end_color='366092',
            fill_type='solid'
        )
        
        # 가운데 정렬 설정
        center_alignment = Alignment(
          horizontal='center',  # 가로 가운데 정렬
          vertical='center'     # 세로 가운데 정렬
        )
        
        # 테두리 설정
        border = Border(
          left=Side(style='thin'),
          right=Side(style='thin'),
          top=Side(style='thin'),
          bottom=Side(style='thin')
        )
        
        # 헤더(첫 번째 행) 스타일 적용
        for cell in ws[1]:
          cell.font = header_font
          cell.fill = header_fill
          cell.alignment = center_alignment
          cell.border = border
        
        # 데이터 셀 스타일 적용
        for row in ws.iter_rows(min_row=2):  # 두 번째 행부터
          for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')  # 세로만 가운데 정렬
      
        # 열 너비 자동 조절
        for column in ws.columns:
          max_length = 0
          column_letter = get_column_letter(column[0].column)
          
          # 각 열의 최대 길이 계산
          for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
          
          # 열 너비 설정 (최대 길이 + 여유 공간)
          adjusted_width = max_length + 2
          
          # 최소, 최대 너비 제한
          adjusted_width = max(100, min(adjusted_width, 50))
          ws.column_dimensions[column_letter].width = adjusted_width
        
        # 행 높이 설정
        for row in ws.rows:
            ws.row_dimensions[row[0].row].height = 25  # 기본 행 높이
        
        # 첫 번째 행(헤더) 특별 설정
        ws.row_dimensions[1].height = 25  # 헤더 행 높이
        
        # 변경사항 저장
        wb.save(filename)
        print("엑셀 파일 형식이 성공적으로 조정되었습니다.")
        
    except Exception as e:
        print(f"파일 형식 조정 중 오류가 발생했습니다: {str(e)}")

# 엑셀에 데이터를 저장하는 함수
def save_to_excel(news_title_text, find_keyword_list, filename=None):
    # 파일명이 지정되지 않은 경우 현재 날짜로 생성
    if filename is None:
      current_date = datetime.now().strftime('%Y%m%d')
      filename = f'news_data_{current_date}.xlsx'
      
    # 두 리스트의 길이를 맞춰주기
    max_length = max(len(news_title_text), len(find_keyword_list))
    news_title_text = news_title_text + [None] * (max_length - len(news_title_text))
    find_keyword_list = find_keyword_list + [None] * (max_length - len(find_keyword_list))  

    # 엑셀에 데이터를 저장하기 위해서는 딕셔너리가 필수!  
    news_list = {
      '뉴스 제목': news_title_text,
      '키워드 뉴스 제목' : find_keyword_list
    }        
    
    try:
      # 데이터프레임 생성(추출한 데이터를 엑셀에 저장)
      df = pd.DataFrame(news_list)
      
      excel_file = f"C:\work\python_projects\scrap_data\{filename}"
      
      # 엑셀 파일로 저장
      df.to_excel(excel_file, index=False, engine='openpyxl')
      
      # 엑셀 형식 조정
      adjust_excel_format(excel_file)
      
      print(f"데이터가 성공적으로 {filename}에 저장되었습니다.")
      
      # 기본적인 데이터 통계 출력
      print("\n데이터 통계:")
      print(f"총 기사 수: {len(df)}")
        
    except Exception as e:
        print(f"엑셀 저장 중 오류 발생: {e}")  

# 엑셀에 데이터 저장 함수 실행
save_to_excel(news_title_text, find_keyword_list)